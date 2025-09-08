from fastapi import FastAPI, Depends, HTTPException, status, Query, Response, Request
from starlette.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import (
    text,
    inspect,
    select,
    func,
    and_,
    union_all,
    literal_column,
    case,
    DateTime,
    Date,
)
from sqlalchemy.sql.schema import Column
from database import (
    get_db,
    create_tables,
    engine,
    SessionLocal,
    User as DBUser,
    Branch as DBBranch,
    Medicine as DBMedicine,
    Employee as DBEmployee,
    Patient as DBPatient,
    Transfer as DBTransfer,
    DispensingRecord as DBDispensingRecord,
    DispensingItem as DBDispensingItem,
    Arrival as DBArrival,
    Category as DBCategory,
    MedicalDevice as DBMedicalDevice,
    Shipment as DBShipment,
    ShipmentItem as DBShipmentItem,
    Notification as DBNotification,
)
from schemas import *
from typing import List, Optional, Iterable, Callable
from datetime import datetime, date, timedelta, timezone, time
from zoneinfo import ZoneInfo
import uuid
import json
from pydantic import ValidationError
from services.stock import get_available_qty, decrement_stock, ItemType
import traceback
import logging
import re
from io import BytesIO
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Workbook = None
    get_column_letter = None

logger = logging.getLogger(__name__)
log = logging.getLogger("reports")

ALMATY_TZ = ZoneInfo("Asia/Almaty")
MAIN_BRANCH_ID = None


def to_almaty(dt: datetime | str | None) -> str:
    if not dt:
        return ""
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return dt
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("UTC"))
    return dt.astimezone(ALMATY_TZ).strftime("%d.%m.%Y %H:%M")


def humanize_items(raw) -> str:
    data = raw
    if isinstance(raw, str):
        try:
            data = json.loads(raw)
        except Exception:
            return raw
    if not isinstance(data, list):
        return ""
    parts = []
    for it in data:
        name = (
            it.get("name")
            or it.get("medicine_name")
            or it.get("device_name")
            or ""
        )
        qty = it.get("quantity") or 0
        parts.append(f"{name} — {qty} шт.")
    return "; ".join(parts)

def ensure_schema_patches():
    with engine.begin() as conn:
        insp = inspect(conn)

        cols_meds = [c["name"] for c in insp.get_columns("medicines")]
        if "category_id" not in cols_meds:
            conn.exec_driver_sql(
                "ALTER TABLE public.medicines ADD COLUMN category_id varchar"
            )

        conn.exec_driver_sql(
            """
        DO $$
        BEGIN
          IF EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'fk_medicines_category') THEN
            ALTER TABLE public.medicines DROP CONSTRAINT fk_medicines_category;
          END IF;
        END $$;
        """
        )

        conn.exec_driver_sql(
            """
        ALTER TABLE public.medicines
          ADD CONSTRAINT fk_medicines_category
          FOREIGN KEY (category_id) REFERENCES public.categories(id) ON DELETE SET NULL;
        """
        )

        conn.exec_driver_sql(
            """
        ALTER TABLE public.medical_devices
          DROP CONSTRAINT IF EXISTS medical_devices_category_id_fkey;
        """
        )
        conn.exec_driver_sql(
            """
        ALTER TABLE public.medical_devices
          DROP CONSTRAINT IF EXISTS fk_medical_devices_category;
        """
        )

        conn.exec_driver_sql(
            """
        ALTER TABLE public.medical_devices
          ADD CONSTRAINT fk_medical_devices_category
          FOREIGN KEY (category_id) REFERENCES public.categories(id) ON DELETE RESTRICT;
        """
        )

        conn.exec_driver_sql(
            "CREATE INDEX IF NOT EXISTS idx_categories_type ON public.categories(type);"
        )
        conn.exec_driver_sql(
            "CREATE INDEX IF NOT EXISTS idx_med_category_id ON public.medicines(category_id);"
        )
        conn.exec_driver_sql(
            "CREATE INDEX IF NOT EXISTS idx_dev_category_id ON public.medical_devices(category_id);"
        )

def ensure_medicines_category_fk():
    inspector = inspect(engine)
    columns = [c["name"] for c in inspector.get_columns("medicines")]
    fks = [fk["name"] for fk in inspector.get_foreign_keys("medicines")]

    with engine.begin() as conn:
        if "category_id" not in columns:
            conn.exec_driver_sql("ALTER TABLE medicines ADD COLUMN category_id VARCHAR")
        if "fk_medicines_category" not in fks:
            conn.exec_driver_sql(
                "ALTER TABLE medicines ADD CONSTRAINT fk_medicines_category FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE SET NULL"
            )
        missing = conn.exec_driver_sql(
            "SELECT 1 FROM medicines WHERE category_id IS NULL LIMIT 1"
        ).first()
        if missing:
            cat = conn.exec_driver_sql(
                "SELECT id FROM categories WHERE type='medicine' LIMIT 1"
            ).first()
            if cat:
                conn.exec_driver_sql(
                    "UPDATE medicines SET category_id = :cid WHERE category_id IS NULL",
                    {"cid": cat[0]},
                )


def ensure_arrivals_schema():
    with engine.begin() as conn:
        insp = inspect(conn)
        cols = {c["name"] for c in insp.get_columns("arrivals")}

        if "item_type" not in cols:
            conn.exec_driver_sql("ALTER TABLE arrivals ADD COLUMN item_type varchar")
        if "item_id" not in cols:
            conn.exec_driver_sql("ALTER TABLE arrivals ADD COLUMN item_id varchar")
        if "item_name" not in cols:
            conn.exec_driver_sql("ALTER TABLE arrivals ADD COLUMN item_name varchar")

        # prices are not part of arrivals anymore
        if "purchase_price" in cols:
            conn.exec_driver_sql("ALTER TABLE arrivals DROP COLUMN purchase_price")
        if "sell_price" in cols:
            conn.exec_driver_sql("ALTER TABLE arrivals DROP COLUMN sell_price")

        # legacy columns from old implementation
        if "medicine_id" in cols:
            conn.exec_driver_sql("ALTER TABLE arrivals DROP COLUMN medicine_id")
        if "medicine_name" in cols:
            conn.exec_driver_sql("ALTER TABLE arrivals DROP COLUMN medicine_name")


# Create FastAPI app
app = FastAPI(title="Warehouse Management System")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create tables on startup
@app.on_event("startup")
async def startup_event():
    create_tables()
    # Create default admin user if not exists
    db = next(get_db())
    admin_user = db.query(DBUser).filter(DBUser.login == "admin").first()
    if not admin_user:
        admin_user = DBUser(
            id="admin",
            login="admin",
            password="admin",
            role="admin"
        )
        db.add(admin_user)
        db.commit()
    
    # Create default categories
    medicine_category = db.query(DBCategory).filter(DBCategory.name == "Общие лекарства").first()
    if not medicine_category:
        medicine_category = DBCategory(
            id=str(uuid.uuid4()),
            name="Общие лекарства",
            description="Общая категория лекарств",
            type="medicine"
        )
        db.add(medicine_category)
    
    device_category = db.query(DBCategory).filter(DBCategory.name == "Общие ИМН").first()
    if not device_category:
        device_category = DBCategory(
            id=str(uuid.uuid4()),
            name="Общие ИМН", 
            description="Общая категория изделий медицинского назначения",
            type="medical_device"
        )
        db.add(device_category)
        
    db.commit()

    ensure_medicines_category_fk()
    ensure_schema_patches()
    ensure_arrivals_schema()
    # Ensure all existing medicines have a valid category and enforce NOT NULL constraint
    try:
        db.execute(
            text("UPDATE medicines SET category_id = :cat WHERE category_id IS NULL"),
            {"cat": medicine_category.id},
        )
        db.commit()
        db.execute(text("ALTER TABLE medicines ALTER COLUMN category_id SET NOT NULL"))
        db.commit()
    except Exception:
        db.rollback()

# Auth endpoints
@app.post("/auth/login", response_model=LoginResponse)
async def login(login_data: UserLogin, db: Session = Depends(get_db)):
    # Check user login
    db_user = db.query(DBUser).filter(DBUser.login == login_data.login, DBUser.password == login_data.password).first()
    if db_user:
        user = User.model_validate(db_user)
        return LoginResponse(
            user=user,
            token=f"token_{db_user.id}"
        )
    
    raise HTTPException(status_code=401, detail="Invalid credentials")

# User endpoints
@app.get("/users", response_model=List[User])
async def get_users(db: Session = Depends(get_db)):
    db_users = db.query(DBUser).all()
    return [User.model_validate(user) for user in db_users]

@app.post("/users", response_model=User)
async def create_user(user: UserCreate, db: Session = Depends(get_db)):
    # Check if user already exists
    existing_user = db.query(DBUser).filter(DBUser.login == user.login).first()
    if existing_user:
        raise HTTPException(status_code=400, detail="User with this login already exists")
    
    user_id = str(uuid.uuid4())
    db_user = DBUser(
        id=user_id,
        login=user.login,
        password=user.password,
        role=user.role,
        branch_name=user.branch_name
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return User.model_validate(db_user)

@app.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user: UserUpdate, db: Session = Depends(get_db)):
    db_user = db.query(DBUser).filter(DBUser.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    for field, value in user.model_dump(exclude_unset=True).items():
        setattr(db_user, field, value)
    
    db.commit()
    db.refresh(db_user)
    return User.model_validate(db_user)

@app.delete("/users/{user_id}")
async def delete_user(user_id: str, db: Session = Depends(get_db)):
    user = db.query(DBUser).filter(DBUser.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    db.delete(user)
    db.commit()
    return {"message": "User deleted"}

# Branch endpoints
@app.get("/branches", response_model=List[Branch])
async def get_branches(db: Session = Depends(get_db)):
    db_branches = db.query(DBBranch).all()
    return [Branch.model_validate(branch) for branch in db_branches]

@app.post("/branches", response_model=Branch)
async def create_branch(branch: BranchCreate, db: Session = Depends(get_db)):
    branch_id = str(uuid.uuid4())
    db_branch = DBBranch(
        id=branch_id,
        name=branch.name,
        login=branch.login,
        password=branch.password
    )
    db.add(db_branch)
    
    # Also create user for branch
    db_user = DBUser(
        id=branch_id,
        login=branch.login,
        password=branch.password,
        role="branch",
        branch_name=branch.name
    )
    db.add(db_user)
    
    db.commit()
    db.refresh(db_branch)
    return Branch.model_validate(db_branch)

@app.put("/branches/{branch_id}", response_model=Branch)
async def update_branch(branch_id: str, branch: BranchUpdate, db: Session = Depends(get_db)):
    db_branch = db.query(DBBranch).filter(DBBranch.id == branch_id).first()
    if not db_branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    for field, value in branch.model_dump(exclude_unset=True).items():
        setattr(db_branch, field, value)
    
    # Update corresponding user
    db_user = db.query(DBUser).filter(DBUser.id == branch_id).first()
    if db_user:
        if branch.login:
            db_user.login = branch.login
        if branch.password:
            db_user.password = branch.password
        if branch.name:
            db_user.branch_name = branch.name
    
    db.commit()
    db.refresh(db_branch)
    return Branch.model_validate(db_branch)

@app.delete("/branches/{branch_id}")
async def delete_branch(branch_id: str, db: Session = Depends(get_db)):
    branch = db.query(DBBranch).filter(DBBranch.id == branch_id).first()
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    # Delete corresponding user
    user = db.query(DBUser).filter(DBUser.id == branch_id).first()
    if user:
        db.delete(user)
    
    db.delete(branch)
    db.commit()
    return {"message": "Branch deleted"}

# Medicine endpoints
@app.get("/medicines", response_model=List[Medicine])
async def get_medicines(branch_id: Optional[str] = None, db: Session = Depends(get_db)):
    if branch_id and branch_id != "null" and branch_id != "undefined":
        db_medicines = db.query(DBMedicine).filter(DBMedicine.branch_id == branch_id).all()
    else:
        db_medicines = db.query(DBMedicine).filter(DBMedicine.branch_id.is_(None)).all()
    return [Medicine.model_validate(medicine) for medicine in db_medicines]

@app.post("/medicines", response_model=Medicine)
async def create_medicine(medicine: MedicineCreate, db: Session = Depends(get_db)):
    cat = db.query(DBCategory).filter(DBCategory.id == medicine.category_id).first()
    if not cat:
        raise HTTPException(status_code=400, detail="Category not found")
    if cat.type != "medicine":
        raise HTTPException(status_code=400, detail="Invalid category for medicine")

    medicine_id = str(uuid.uuid4())
    db_medicine = DBMedicine(
        id=medicine_id,
        name=medicine.name,
        category_id=medicine.category_id,
        purchase_price=medicine.purchase_price,
        sell_price=medicine.sell_price,
        quantity=medicine.quantity,
        branch_id=medicine.branch_id,
    )
    db.add(db_medicine)
    db.commit()
    db.refresh(db_medicine)
    return Medicine.model_validate(db_medicine)

@app.put("/medicines/{medicine_id}", response_model=Medicine)
async def update_medicine(medicine_id: str, medicine: MedicineUpdate, db: Session = Depends(get_db)):
    db_medicine = db.query(DBMedicine).filter(DBMedicine.id == medicine_id).first()
    if not db_medicine:
        raise HTTPException(status_code=404, detail="Medicine not found")

    category_id = medicine.category_id if medicine.category_id is not None else db_medicine.category_id
    cat = db.query(DBCategory).filter(DBCategory.id == medicine.category_id).first()
    if not cat:
        raise HTTPException(status_code=400, detail="Category not found")
    if cat.type != "medicine":
        raise HTTPException(status_code=400, detail="Invalid category for medicine")
    
    for field, value in medicine.model_dump(exclude_unset=True).items():
        setattr(db_medicine, field, value)

    db.commit()
    db.refresh(db_medicine)
    return Medicine.model_validate(db_medicine)

@app.delete("/medicines/{medicine_id}")
async def delete_medicine(medicine_id: str, db: Session = Depends(get_db)):
    medicine = db.query(DBMedicine).filter(DBMedicine.id == medicine_id).first()
    if not medicine:
        raise HTTPException(status_code=404, detail="Medicine not found")
    
    db.delete(medicine)
    db.commit()
    return {"message": "Medicine deleted"}

# Medical Device endpoints
@app.get("/medical_devices", response_model=List[MedicalDevice])
async def get_medical_devices(branch_id: Optional[str] = None, db: Session = Depends(get_db)):
    if branch_id and branch_id != "null" and branch_id != "undefined":
        db_devices = db.query(DBMedicalDevice).filter(DBMedicalDevice.branch_id == branch_id).all()
    else:
        db_devices = db.query(DBMedicalDevice).filter(DBMedicalDevice.branch_id.is_(None)).all()
    return [MedicalDevice.model_validate(device) for device in db_devices]

@app.post("/medical_devices", response_model=MedicalDevice)
async def create_medical_device(device: MedicalDeviceCreate, db: Session = Depends(get_db)):
    cat = db.query(DBCategory).filter(DBCategory.id == device.category_id).first()
    if not cat:
        raise HTTPException(status_code=400, detail="Category not found")
    if cat.type != "medical_device":
        raise HTTPException(status_code=400, detail="Invalid category for medical device")

    device_id = str(uuid.uuid4())
    db_device = DBMedicalDevice(
        id=device_id,
        name=device.name,
        category_id=device.category_id,
        purchase_price=device.purchase_price,
        sell_price=device.sell_price,
        quantity=device.quantity,
        branch_id=device.branch_id,
    )
    db.add(db_device)
    db.commit()
    db.refresh(db_device)
    return MedicalDevice.model_validate(db_device)

@app.put("/medical_devices/{device_id}", response_model=MedicalDevice)
async def update_medical_device(device_id: str, device: MedicalDeviceUpdate, db: Session = Depends(get_db)):
    db_device = db.query(DBMedicalDevice).filter(DBMedicalDevice.id == device_id).first()
    if not db_device:
        raise HTTPException(status_code=404, detail="Medical device not found")

    category_id = device.category_id if device.category_id is not None else db_device.category_id
    cat = db.query(DBCategory).filter(DBCategory.id == device.category_id).first()
    if not cat:
        raise HTTPException(status_code=400, detail="Category not found")
    if cat.type != "medical_device":
        raise HTTPException(status_code=400, detail="Invalid category for medical device")

    for field, value in device.model_dump(exclude_unset=True).items():
        setattr(db_device, field, value)

    db.commit()
    db.refresh(db_device)
    return MedicalDevice.model_validate(db_device)

@app.delete("/medical_devices/{device_id}")
async def delete_medical_device(device_id: str, db: Session = Depends(get_db)):
    device = db.query(DBMedicalDevice).filter(DBMedicalDevice.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Medical device not found")
    
    db.delete(device)
    db.commit()
    return {"message": "Medical device deleted"}

# Category endpoints
@app.get("/categories", response_model=List[dict])
async def get_categories(type: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(DBCategory)
    if type:
        query = query.filter(DBCategory.type == type)
    categories = query.all()
    return [{"id": cat.id, "name": cat.name, "description": cat.description, "type": cat.type} for cat in categories]

@app.post("/categories")
async def create_category(category: dict, db: Session = Depends(get_db)):
    category_id = str(uuid.uuid4())
    db_category = DBCategory(
        id=category_id,
        name=category["name"],
        description=category.get("description"),
        type=category["type"]
    )
    db.add(db_category)
    db.commit()
    db.refresh(db_category)
    return {"id": db_category.id, "name": db_category.name, "description": db_category.description, "type": db_category.type}

@app.put("/categories/{category_id}")
async def update_category(category_id: str, category: dict, db: Session = Depends(get_db)):
    db_category = db.query(DBCategory).filter(DBCategory.id == category_id).first()
    if not db_category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    for field, value in category.items():
        if hasattr(db_category, field):
            setattr(db_category, field, value)
    
    db.commit()
    db.refresh(db_category)
    return {"id": db_category.id, "name": db_category.name, "description": db_category.description, "type": db_category.type}

@app.delete("/categories/{category_id}")
async def delete_category(category_id: str, db: Session = Depends(get_db)):
    category = db.query(DBCategory).filter(DBCategory.id == category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    db.delete(category)
    db.commit()
    return {"message": "Category deleted"}

# Employee endpoints
@app.get("/employees", response_model=List[Employee])
async def get_employees(branch_id: Optional[str] = None, db: Session = Depends(get_db)):
    if branch_id and branch_id != "null" and branch_id != "undefined":
        db_employees = db.query(DBEmployee).filter(DBEmployee.branch_id == branch_id).all()
    else:
        db_employees = db.query(DBEmployee).filter(DBEmployee.branch_id.is_(None)).all()
    return [Employee.model_validate(employee) for employee in db_employees]

@app.post("/employees", response_model=Employee)
async def create_employee(employee: EmployeeCreate, db: Session = Depends(get_db)):
    employee_id = str(uuid.uuid4())
    db_employee = DBEmployee(
        id=employee_id,
        first_name=employee.first_name,
        last_name=employee.last_name,
        phone=employee.phone,
        address=employee.address,
        branch_id=employee.branch_id
    )
    db.add(db_employee)
    db.commit()
    db.refresh(db_employee)
    return Employee.model_validate(db_employee)

@app.put("/employees/{employee_id}", response_model=Employee)
async def update_employee(employee_id: str, employee: EmployeeUpdate, db: Session = Depends(get_db)):
    db_employee = db.query(DBEmployee).filter(DBEmployee.id == employee_id).first()
    if not db_employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    for field, value in employee.model_dump(exclude_unset=True).items():
        setattr(db_employee, field, value)
    
    db.commit()
    db.refresh(db_employee)
    return Employee.model_validate(db_employee)

@app.delete("/employees/{employee_id}")
async def delete_employee(employee_id: str, db: Session = Depends(get_db)):
    employee = db.query(DBEmployee).filter(DBEmployee.id == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    db.delete(employee)
    db.commit()
    return {"message": "Employee deleted"}

# Patient endpoints
@app.get("/patients", response_model=List[Patient])
async def get_patients(branch_id: Optional[str] = None, db: Session = Depends(get_db)):
    if branch_id and branch_id != "null" and branch_id != "undefined":
        db_patients = db.query(DBPatient).filter(DBPatient.branch_id == branch_id).all()
    else:
        db_patients = db.query(DBPatient).all()
    return [Patient.model_validate(patient) for patient in db_patients]

@app.post("/patients", response_model=Patient)
async def create_patient(patient: PatientCreate, db: Session = Depends(get_db)):
    patient_id = str(uuid.uuid4())
    db_patient = DBPatient(
        id=patient_id,
        first_name=patient.first_name,
        last_name=patient.last_name,
        illness=patient.illness,
        phone=patient.phone,
        address=patient.address,
        branch_id=patient.branch_id
    )
    db.add(db_patient)
    db.commit()
    db.refresh(db_patient)
    return Patient.model_validate(db_patient)

@app.put("/patients/{patient_id}", response_model=Patient)
async def update_patient(patient_id: str, patient: PatientUpdate, db: Session = Depends(get_db)):
    db_patient = db.query(DBPatient).filter(DBPatient.id == patient_id).first()
    if not db_patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    for field, value in patient.model_dump(exclude_unset=True).items():
        setattr(db_patient, field, value)
    
    db.commit()
    db.refresh(db_patient)
    return Patient.model_validate(db_patient)

@app.delete("/patients/{patient_id}")
async def delete_patient(patient_id: str, db: Session = Depends(get_db)):
    patient = db.query(DBPatient).filter(DBPatient.id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    db.delete(patient)
    db.commit()
    return {"message": "Patient deleted"}

# Transfer endpoints
@app.get("/transfers", response_model=List[Transfer])
async def get_transfers(branch_id: Optional[str] = None, db: Session = Depends(get_db)):
    if branch_id and branch_id != "null" and branch_id != "undefined":
        db_transfers = db.query(DBTransfer).filter(DBTransfer.to_branch_id == branch_id).all()
    else:
        db_transfers = db.query(DBTransfer).all()
    return [Transfer.model_validate(transfer) for transfer in db_transfers]

@app.post("/transfers")
async def create_transfers(batch: BatchTransferCreate, db: Session = Depends(get_db)):
    try:
        for transfer_data in batch.transfers:
            # Check main warehouse medicine
            main_medicine = db.query(DBMedicine).filter(
                DBMedicine.id == transfer_data.medicine_id,
                DBMedicine.branch_id.is_(None)
            ).first()
            
            if not main_medicine or main_medicine.quantity < transfer_data.quantity:
                raise HTTPException(status_code=400, detail=f"Not enough {transfer_data.medicine_name} in main warehouse")
            
            # Decrease quantity in main warehouse
            main_medicine.quantity -= transfer_data.quantity
            
            # Find or create medicine in branch
            branch_medicine = db.query(DBMedicine).filter(
                DBMedicine.name == transfer_data.medicine_name,
                DBMedicine.branch_id == transfer_data.to_branch_id
            ).first()
            
            if branch_medicine:
                branch_medicine.quantity += transfer_data.quantity
            else:
                new_branch_medicine = DBMedicine(
                    id=str(uuid.uuid4()),
                    name=transfer_data.medicine_name,
                    category_id=main_medicine.category_id,
                    purchase_price=main_medicine.purchase_price,
                    sell_price=main_medicine.sell_price,
                    quantity=transfer_data.quantity,
                    branch_id=transfer_data.to_branch_id
                )
                db.add(new_branch_medicine)
            
            # Create transfer record
            db_transfer = DBTransfer(
                id=str(uuid.uuid4()),
                medicine_id=transfer_data.medicine_id,
                medicine_name=transfer_data.medicine_name,
                quantity=transfer_data.quantity,
                from_branch_id=transfer_data.from_branch_id or "main",
                to_branch_id=transfer_data.to_branch_id
            )
            db.add(db_transfer)
        
        db.commit()
        return {"message": "Transfers completed"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# Shipment endpoints
@app.get("/shipments")
async def get_shipments(branch_id: Optional[str] = None, db: Session = Depends(get_db)):
    if branch_id and branch_id != "null" and branch_id != "undefined":
        shipments = db.query(DBShipment).filter(DBShipment.to_branch_id == branch_id).all()
    else:
        shipments = db.query(DBShipment).all()
    
    result = []
    for shipment in shipments:
        # Get shipment items
        items = db.query(DBShipmentItem).filter(DBShipmentItem.shipment_id == shipment.id).all()
        
        shipment_data = {
            "id": shipment.id,
            "to_branch_id": shipment.to_branch_id,
            "status": shipment.status,
            "rejection_reason": shipment.rejection_reason,
            "created_at": shipment.created_at.isoformat(),
            "medicines": [],
            "medical_devices": []
        }
        
        for item in items:
            if item.item_type == "medicine":
                shipment_data["medicines"].append({
                    "medicine_id": item.item_id,
                    "name": item.item_name,
                    "quantity": item.quantity
                })
            else:
                shipment_data["medical_devices"].append({
                    "device_id": item.item_id,
                    "name": item.item_name,
                    "quantity": item.quantity
                })
        
        result.append(shipment_data)
    
    return {"data": result}

@app.get("/dispensing_records/{record_id}")
async def get_dispensing_record_detail(record_id: str, db: Session = Depends(get_db)):
    record = (
        db.query(DBDispensingRecord, DBBranch)
        .join(DBBranch, DBDispensingRecord.branch_id == DBBranch.id)
        .filter(DBDispensingRecord.id == record_id)
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    rec, branch = record
    items = db.query(DBDispensingItem).filter(DBDispensingItem.record_id == record_id).all()
    tz = ZoneInfo("Asia/Almaty")
    dt = rec.date
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    local_dt = dt.astimezone(tz)
    data = {
        "id": rec.id,
        "time": local_dt.strftime("%H:%M:%S"),
        "patient_name": rec.patient_name,
        "employee_name": rec.employee_name,
        "branch_name": branch.name if branch else "",
        "items": [
            {
                "type": item.item_type,
                "name": item.item_name,
                "quantity": item.quantity,
            }
            for item in items
        ],
    }
    return {"data": data}

@app.post("/shipments")
async def create_shipment(shipment_data: dict, db: Session = Depends(get_db)):
    try:
        shipment_id = str(uuid.uuid4())
        
        # Create shipment
        db_shipment = DBShipment(
            id=shipment_id,
            to_branch_id=shipment_data["to_branch_id"],
            status="pending"
        )
        db.add(db_shipment)
        
        # Add medicines
        if "medicines" in shipment_data:
            for medicine_item in shipment_data["medicines"]:
                medicine = db.query(DBMedicine).filter(
                    DBMedicine.id == medicine_item["medicine_id"],
                    DBMedicine.branch_id.is_(None)
                ).first()
                
                if not medicine or medicine.quantity < medicine_item["quantity"]:
                    raise HTTPException(status_code=400, detail=f"Insufficient medicine quantity")
                
                # Create shipment item
                db_item = DBShipmentItem(
                    id=str(uuid.uuid4()),
                    shipment_id=shipment_id,
                    item_type="medicine",
                    item_id=medicine_item["medicine_id"],
                    item_name=medicine.name,
                    quantity=medicine_item["quantity"]
                )
                db.add(db_item)
        
        # Add medical devices
        if "medical_devices" in shipment_data:
            for device_item in shipment_data["medical_devices"]:
                device = db.query(DBMedicalDevice).filter(
                    DBMedicalDevice.id == device_item["device_id"],
                    DBMedicalDevice.branch_id.is_(None)
                ).first()
                
                if not device or device.quantity < device_item["quantity"]:
                    raise HTTPException(status_code=400, detail=f"Insufficient medical device quantity")
                
                # Create shipment item
                db_item = DBShipmentItem(
                    id=str(uuid.uuid4()),
                    shipment_id=shipment_id,
                    item_type="medical_device",
                    item_id=device_item["device_id"],
                    item_name=device.name,
                    quantity=device_item["quantity"]
                )
                db.add(db_item)
        
        # Create notification for branch
        notification = DBNotification(
            id=str(uuid.uuid4()),
            branch_id=shipment_data["to_branch_id"],
            title="Новая отправка",
            message=f"Поступление от главного склада",
            is_read=0
        )
        db.add(notification)
        
        db.commit()
        return {"message": "Shipment created successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/shipments/{shipment_id}/accept")
async def accept_shipment(shipment_id: str, db: Session = Depends(get_db)):
    try:
        shipment = db.query(DBShipment).filter(DBShipment.id == shipment_id).first()
        if not shipment:
            raise HTTPException(status_code=404, detail="Shipment not found")
        
        # Get shipment items
        items = db.query(DBShipmentItem).filter(DBShipmentItem.shipment_id == shipment_id).all()
        
        for item in items:
            if item.item_type == "medicine":
                # Decrease main warehouse quantity
                main_medicine = db.query(DBMedicine).filter(
                    DBMedicine.id == item.item_id,
                    DBMedicine.branch_id.is_(None)
                ).first()
                if main_medicine:
                    main_medicine.quantity -= item.quantity
                
                # Add to branch
                branch_medicine = db.query(DBMedicine).filter(
                    DBMedicine.name == item.item_name,
                    DBMedicine.branch_id == shipment.to_branch_id
                ).first()
                
                if branch_medicine:
                    branch_medicine.quantity += item.quantity
                else:
                    new_medicine = DBMedicine(
                        id=str(uuid.uuid4()),
                        name=item.item_name,
                        category_id=main_medicine.category_id if main_medicine else None,
                        purchase_price=main_medicine.purchase_price if main_medicine else 0,
                        sell_price=main_medicine.sell_price if main_medicine else 0,
                        quantity=item.quantity,
                        branch_id=shipment.to_branch_id
                    )
                    db.add(new_medicine)
            
            elif item.item_type == "medical_device":
                # Decrease main warehouse quantity
                main_device = db.query(DBMedicalDevice).filter(
                    DBMedicalDevice.id == item.item_id,
                    DBMedicalDevice.branch_id.is_(None)
                ).first()
                if main_device:
                    main_device.quantity -= item.quantity
                
                # Add to branch
                branch_device = db.query(DBMedicalDevice).filter(
                    DBMedicalDevice.name == item.item_name,
                    DBMedicalDevice.branch_id == shipment.to_branch_id
                ).first()
                
                if branch_device:
                    branch_device.quantity += item.quantity
                else:
                    new_device = DBMedicalDevice(
                        id=str(uuid.uuid4()),
                        name=item.item_name,
                        category_id=main_device.category_id if main_device else None,
                        purchase_price=main_device.purchase_price if main_device else 0,
                        sell_price=main_device.sell_price if main_device else 0,
                        quantity=item.quantity,
                        branch_id=shipment.to_branch_id
                    )
                    db.add(new_device)
        
        shipment.status = "accepted"
        db.commit()
        return {"message": "Shipment accepted"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/shipments/{shipment_id}/reject")
async def reject_shipment(shipment_id: str, reason: dict, db: Session = Depends(get_db)):
    shipment = db.query(DBShipment).filter(DBShipment.id == shipment_id).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    shipment.status = "rejected"
    shipment.rejection_reason = reason.get("reason", "")
    db.commit()
    return {"message": "Shipment rejected"}

@app.put("/shipments/{shipment_id}/status")
async def update_shipment_status(shipment_id: str, status_data: dict, db: Session = Depends(get_db)):
    shipment = db.query(DBShipment).filter(DBShipment.id == shipment_id).first()
    if not shipment:
        raise HTTPException(status_code=404, detail="Shipment not found")
    
    shipment.status = status_data["status"]
    db.commit()
    return {"message": "Shipment status updated"}

# Notification endpoints
@app.get("/notifications")
async def get_notifications(branch_id: Optional[str] = None, db: Session = Depends(get_db)):
    if branch_id:
        notifications = db.query(DBNotification).filter(DBNotification.branch_id == branch_id).order_by(DBNotification.created_at.desc()).all()
    else:
        notifications = db.query(DBNotification).order_by(DBNotification.created_at.desc()).all()
    
    return {
        "data": [
            {
                "id": n.id,
                "branch_id": n.branch_id,
                "title": n.title,
                "message": n.message,
                "is_read": bool(n.is_read),
                "created_at": n.created_at.isoformat()
            }
            for n in notifications
        ]
    }

@app.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, db: Session = Depends(get_db)):
    notification = db.query(DBNotification).filter(DBNotification.id == notification_id).first()
    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")
    
    notification.is_read = 1
    db.commit()
    return {"message": "Notification marked as read"}

# Dispensing endpoints
@app.get("/dispensing_records")
async def get_dispensing_records(branch_id: Optional[str] = None, db: Session = Depends(get_db)):
    if branch_id and branch_id != "null" and branch_id != "undefined":
        records = db.query(DBDispensingRecord).filter(DBDispensingRecord.branch_id == branch_id).all()
    else:
        records = db.query(DBDispensingRecord).all()
    
    result = []
    for record in records:
        items = db.query(DBDispensingItem).filter(DBDispensingItem.record_id == record.id).all()
        
        record_data = {
            "id": record.id,
            "patient_id": record.patient_id,
            "patient_name": record.patient_name,
            "employee_id": record.employee_id,
            "employee_name": record.employee_name,
            "branch_id": record.branch_id,
            "date": record.date.isoformat(),
            "medicines": [],
            "medical_devices": []
        }
        
        for item in items:
            if item.item_type == "medicine":
                record_data["medicines"].append({
                    "medicine_name": item.item_name,
                    "quantity": item.quantity
                })
            else:
                record_data["medical_devices"].append({
                    "device_name": item.item_name,
                    "quantity": item.quantity
                })
        
        result.append(record_data)
    
    return {"data": result}

@app.post("/dispensing", status_code=status.HTTP_201_CREATED)
async def create_dispensing_record(payload: dict, db: Session = Depends(get_db)):
    print("DISPENSING_RAW", payload)
    try:
        body = DispensePayload.model_validate(payload)
    except ValidationError as e:
        raise HTTPException(status_code=400, detail=e.errors())

    # Normalize items to unified shape and convert types
    raw_items: list[DispenseLine] = body._normalized_items
    items = [
        {
            "type": ItemType(line.item_type),
            "item_id": line.item_id,
            "quantity": int(line.quantity),
        }
        for line in raw_items
        if line.quantity > 0
    ]
    print("DISPENSING_ITEMS", items)
    if not items:
        raise HTTPException(status_code=400, detail="No items to dispense")

    def parse_uuid_or_str(val, use_uuid: bool):
        if use_uuid:
            try:
                return uuid.UUID(str(val))
            except Exception:
                raise HTTPException(status_code=400, detail=f"Invalid UUID: {val}")
        return str(val)

    try:
        ids_are_uuid = (
            getattr(DBDispensingRecord.__table__.c.id.type.python_type, "__name__", "")
            == "UUID"
        )
        branch_id = parse_uuid_or_str(body.branch_id, ids_are_uuid)
        patient_id = parse_uuid_or_str(body.patient_id, ids_are_uuid)
        employee_id = parse_uuid_or_str(body.employee_id, ids_are_uuid)
        for itm in items:
            itm["item_id"] = parse_uuid_or_str(itm["item_id"], ids_are_uuid)

        print(
            "DISPENSING_IDS",
            branch_id,
            patient_id,
            employee_id,
        )

        with db.begin():
            patient = db.query(DBPatient).filter(DBPatient.id == str(patient_id)).first()
            employee = db.query(DBEmployee).filter(DBEmployee.id == str(employee_id)).first()
            if not patient or not employee:
                raise HTTPException(status_code=404, detail="Patient or employee not found")

            for itm in items:
                available, name = get_available_qty(
                    db, str(branch_id), itm["type"], str(itm["item_id"])
                )
                if itm["quantity"] > available:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"Not enough stock for {itm['type'].value}:{itm['item_id']} (available {available}, requested {itm['quantity']})"
                        ),
                    )
                itm["item_name"] = name

            db_record = DBDispensingRecord(
                id=str(uuid.uuid4()),
                branch_id=str(branch_id),
                patient_id=str(patient_id),
                patient_name=body.patient_name
                or f"{patient.first_name} {patient.last_name}",
                employee_id=str(employee_id),
                employee_name=body.employee_name
                or f"{employee.first_name} {employee.last_name}",
            )
            db.add(db_record)
            db.flush()

            for itm in items:
                db.add(
                    DBDispensingItem(
                        id=str(uuid.uuid4()),
                        record_id=db_record.id,
                        item_type=itm["type"].value,
                        item_id=str(itm["item_id"]),
                        item_name=itm["item_name"],
                        quantity=itm["quantity"],
                    )
                )
                decrement_stock(
                    db,
                    str(branch_id),
                    itm["type"],
                    str(itm["item_id"]),
                    itm["quantity"],
                )

            return {
                "id": db_record.id,
                "branch_id": str(branch_id),
                "patient_id": str(patient_id),
                "employee_id": str(employee_id),
                "items": [
                    {
                        "type": itm["type"].value,
                        "item_id": str(itm["item_id"]),
                        "quantity": itm["quantity"],
                    }
                    for itm in items
                ],
            }
    except HTTPException:
        raise
    except Exception as e:
        print("DISPENSING_ERROR:", repr(e))
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Failed to create dispensing")

# Arrival endpoints
@app.get("/arrivals")
async def get_arrivals(item_type: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(DBArrival)
    if item_type:
        q = q.filter(DBArrival.item_type == item_type)
    rows = q.all()
    return {"data": [Arrival.model_validate(x) for x in rows]}

@app.post("/arrivals")
async def create_arrivals(batch: BatchArrivalCreate, db: Session = Depends(get_db)):
    try:
        for it in batch.arrivals:
            db.add(DBArrival(
                id=str(uuid.uuid4()),
                item_type=it.item_type,
                item_id=it.item_id,
                item_name=it.item_name,
                quantity=it.quantity,
            ))

            # increase stock on MAIN warehouse (branch_id IS NULL)
            if it.item_type == "medicine":
                stock = db.query(DBMedicine).filter(
                    DBMedicine.id == it.item_id,
                    DBMedicine.branch_id.is_(None),
                ).first()
            elif it.item_type == "medical_device":
                stock = db.query(DBMedicalDevice).filter(
                    DBMedicalDevice.id == it.item_id,
                    DBMedicalDevice.branch_id.is_(None),
                ).first()
            else:
                raise HTTPException(status_code=400, detail="Invalid item_type")

            if not stock:
                raise HTTPException(status_code=404, detail="Item not found")

            stock.quantity += it.quantity  # do not modify prices here

        db.commit()
        return {"message": "Arrivals created successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# Report endpoints
@app.post("/reports/generate")
async def generate_report(request: ReportRequest, db: Session = Depends(get_db)):
    try:
        report_data = []
        
        if request.type == "stock":
            if request.branch_id:
                medicines = db.query(DBMedicine).filter(DBMedicine.branch_id == request.branch_id).all()
                devices = db.query(DBMedicalDevice).filter(DBMedicalDevice.branch_id == request.branch_id).all()
            else:
                medicines = db.query(DBMedicine).filter(DBMedicine.branch_id.is_(None)).all()
                devices = db.query(DBMedicalDevice).filter(DBMedicalDevice.branch_id.is_(None)).all()
            
            for med in medicines:
                report_data.append({
                    "id": med.id,
                    "name": med.name,
                    "type": "medicine",
                    "quantity": med.quantity,
                    "purchase_price": med.purchase_price,
                    "sell_price": med.sell_price
                })
            
            for dev in devices:
                report_data.append({
                    "id": dev.id,
                    "name": dev.name,
                    "type": "medical_device",
                    "quantity": dev.quantity,
                    "purchase_price": dev.purchase_price,
                    "sell_price": dev.sell_price
                })
        
        elif request.type == "dispensing":
            query = db.query(DBDispensingRecord)
            if request.branch_id:
                query = query.filter(DBDispensingRecord.branch_id == request.branch_id)
            if request.date_from:
                query = query.filter(DBDispensingRecord.date >= request.date_from)
            if request.date_to:
                query = query.filter(DBDispensingRecord.date <= request.date_to)
            
            records = query.all()
            for record in records:
                report_data.append({
                    "id": record.id,
                    "patient_name": record.patient_name,
                    "employee_name": record.employee_name,
                    "date": record.date.isoformat(),
                    "branch_id": record.branch_id
                })
        
        elif request.type == "arrivals":
            query = db.query(DBArrival)
            if request.date_from:
                query = query.filter(DBArrival.date >= request.date_from)
            if request.date_to:
                query = query.filter(DBArrival.date <= request.date_to)

            arrivals = query.all()
            for arrival in arrivals:
                report_data.append({
                    "id": arrival.id,
                    "item_type": arrival.item_type,
                    "item_name": arrival.item_name,
                    "quantity": arrival.quantity,
                    "date": arrival.date.isoformat()
                })
        
        elif request.type == "transfers":
            query = db.query(DBTransfer)
            if request.branch_id:
                query = query.filter(DBTransfer.to_branch_id == request.branch_id)
            if request.date_from:
                query = query.filter(DBTransfer.date >= request.date_from)
            if request.date_to:
                query = query.filter(DBTransfer.date <= request.date_to)
            
            transfers = query.all()
            for transfer in transfers:
                report_data.append({
                    "id": transfer.id,
                    "medicine_name": transfer.medicine_name,
                    "quantity": transfer.quantity,
                    "from_branch_id": transfer.from_branch_id,
                    "to_branch_id": transfer.to_branch_id,
                    "date": transfer.date.isoformat()
                })
        
        elif request.type == "patients":
            query = db.query(DBPatient)
            if request.branch_id:
                query = query.filter(DBPatient.branch_id == request.branch_id)
            
            patients = query.all()
            for patient in patients:
                report_data.append({
                    "id": patient.id,
                    "first_name": patient.first_name,
                    "last_name": patient.last_name,
                    "illness": patient.illness,
                    "phone": patient.phone,
                    "address": patient.address,
                    "branch_id": patient.branch_id
                })
        
        elif request.type == "medical_devices":
            if request.branch_id:
                devices = db.query(DBMedicalDevice).filter(DBMedicalDevice.branch_id == request.branch_id).all()
            else:
                devices = db.query(DBMedicalDevice).filter(DBMedicalDevice.branch_id.is_(None)).all()
            
            for device in devices:
                report_data.append({
                    "id": device.id,
                    "name": device.name,
                    "quantity": device.quantity,
                    "purchase_price": device.purchase_price,
                    "sell_price": device.sell_price,
                    "branch_id": device.branch_id
                })
        
        return {"data": report_data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/reports/dispensing")
async def get_dispensing_report(
    branch_id: str,
    date_from: str,
    date_to: str,
    db: Session = Depends(get_db),
):
    try:
        start = datetime.fromisoformat(date_from).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        end = datetime.fromisoformat(date_to).replace(
            hour=23, minute=59, second=59, microsecond=999999
        )

        records = (
            db.query(DBDispensingRecord)
            .options(joinedload(DBDispensingRecord.items))
            .filter(
                DBDispensingRecord.branch_id == branch_id,
                DBDispensingRecord.date >= start,
                DBDispensingRecord.date <= end,
            )
            .all()
        )

        data = []
        for r in records:
            items = [
                {
                    "type": i.item_type,
                    "name": i.item_name,
                    "quantity": i.quantity,
                }
                for i in r.items
            ]
            data.append(
                {
                    "id": r.id,
                    "patient_name": r.patient_name,
                    "employee_name": r.employee_name,
                    "datetime": r.date.isoformat(),
                    "items": items,
                }
            )

        return {"data": data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/reports/incoming")
async def get_incoming_report(
    branch_id: str,
    date_from: str,
    date_to: str,
    db: Session = Depends(get_db),
):
    try:
        start = datetime.fromisoformat(date_from).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        end = datetime.fromisoformat(date_to).replace(
            hour=23, minute=59, second=59, microsecond=999999
        )

        shipments = (
            db.query(DBShipment)
            .filter(
                DBShipment.to_branch_id == branch_id,
                DBShipment.status == "accepted",
                DBShipment.created_at >= start,
                DBShipment.created_at <= end,
            )
            .all()
        )

        data = []
        for s in shipments:
            items = (
                db.query(DBShipmentItem)
                .filter(DBShipmentItem.shipment_id == s.id)
                .all()
            )
            items_data = [
                {"type": it.item_type, "name": it.item_name, "quantity": it.quantity}
                for it in items
            ]
            data.append(
                {
                    "id": s.id,
                    "datetime": s.created_at.isoformat(),
                    "items": items_data,
                }
            )

        return {"data": data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/reports/dispensing/export")
async def export_dispensing_report(
    branch_id: str,
    date_from: str,
    date_to: str,
    db: Session = Depends(get_db),
):
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    payload = await get_dispensing_report(
        branch_id=branch_id, date_from=date_from, date_to=date_to, db=db
    )
    records = payload.get("data", [])
    rows = [
        {
            "Пациент": r.get("patient_name", ""),
            "Сотрудник": r.get("employee_name", ""),
            "Дата": to_almaty(r.get("created_at") or r.get("datetime")),
            "Что выдано": humanize_items(
                r.get("items")
                or r.get("medicines")
                or r.get("medical_devices")
            ),
        }
        for r in records
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Выдачи"
    headers = ["Пациент", "Сотрудник", "Дата", "Что выдано"]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "Отчет по выдачам.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/reports/incoming/export")
async def export_incoming_report(
    branch_id: str,
    date_from: str,
    date_to: str,
    db: Session = Depends(get_db),
):
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    payload = await get_incoming_report(
        branch_id=branch_id, date_from=date_from, date_to=date_to, db=db
    )
    records = payload.get("data", [])
    rows = [
        {
            "Дата": to_almaty(r.get("created_at") or r.get("datetime")),
            "Поступило": humanize_items(r.get("items")),
        }
        for r in records
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Поступления"
    headers = ["Дата", "Поступило"]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "Отчет по поступлениям.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/reports/dispensings")
async def get_dispensings_report(
    request: Request,
    date_from: str,
    date_to: str,
    branch_id: str | None = Query(None),
    export: str | None = Query(None),
    format: str | None = Query(None),
    db: Session = Depends(get_db),
):
    start = datetime.fromisoformat(date_from).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    end = datetime.fromisoformat(date_to).replace(
        hour=23, minute=59, second=59, microsecond=999999
    )

    q = db.query(DBDispensingRecord).options(joinedload(DBDispensingRecord.items))
    q = q.filter(DBDispensingRecord.date >= start, DBDispensingRecord.date <= end)
    if branch_id:
        q = q.filter(DBDispensingRecord.branch_id == branch_id)
    records = q.all()

    patient_ids = {r.patient_id for r in records}
    employee_ids = {r.employee_id for r in records}
    name_map = {(i.item_type, i.item_id): i.item_name for r in records for i in r.items}

    patient_map = {
        p.id: f"{p.first_name} {p.last_name}".strip()
        for p in db.query(DBPatient).filter(DBPatient.id.in_(patient_ids)).all()
    } if patient_ids else {}
    employee_map = {
        e.id: f"{e.first_name} {e.last_name}".strip()
        for e in db.query(DBEmployee).filter(DBEmployee.id.in_(employee_ids)).all()
    } if employee_ids else {}

    med_ids = [iid for (t, iid) in name_map if t == "medicine"]
    if med_ids:
        for m in db.query(DBMedicine.id, DBMedicine.name).filter(DBMedicine.id.in_(med_ids)):
            name_map[("medicine", m.id)] = m.name
    dev_ids = [iid for (t, iid) in name_map if t == "medical_device"]
    if dev_ids:
        for d in db.query(DBMedicalDevice.id, DBMedicalDevice.name).filter(
            DBMedicalDevice.id.in_(dev_ids)
        ):
            name_map[("medical_device", d.id)] = d.name

    json_rows = []
    for r in records:
        patient_full = patient_map.get(r.patient_id, r.patient_name or "")
        employee_full = employee_map.get(r.employee_id, r.employee_name or "")
        dt_iso = r.date.isoformat() if r.date else ""

        json_items = []
        for i in r.items:
            nm = name_map.get((i.item_type, i.item_id), i.item_name)
            json_items.append(
                {"type": i.item_type, "name": nm, "quantity": i.quantity}
            )

        json_rows.append(
            {
                "id": r.id,
                "patient_name": patient_full,
                "employee_name": employee_full,
                "datetime": dt_iso,
                "items": json_items,
            }
        )

    result = {"data": json_rows}

    if _wants_excel(export, format):
        rows = []
        for r in result.get("data", []):
            items_list = r.get("items", []) or []
            items_human = "; ".join(
                f"{i.get('name','')} — {i.get('quantity','')}" for i in items_list
            )
            rows.append(
                [
                    r.get("patient_name", ""),
                    r.get("employee_name", ""),
                    _to_almaty_str(r.get("datetime", "")),
                    items_human,
                ]
            )

        content = _render_xlsx(
            headers=[
                "Пациент",
                "Сотрудник",
                "Дата и время",
                "Выдано (наименование — кол-во)",
            ],
            rows=rows,
            sheet_name="Выдачи",
        )

        safe_to = date_to or datetime.now(ALMATY_TZ).strftime("%Y-%m-%d")
        return Response(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=_ascii_headers(
                f"dispensings_report_{safe_to}.xlsx"
            ),
        )

    return result


async def build_arrivals_json_payload(
    *, branch_id: str | None, date_from: str, date_to: str, db: Session
) -> dict:
    start = datetime.fromisoformat(date_from).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    end = datetime.fromisoformat(date_to).replace(
        hour=23, minute=59, second=59, microsecond=999999
    )

    col = pick_arrival_branch_col(DBArrival)
    q = db.query(DBArrival)
    q = q.filter(DBArrival.date >= start, DBArrival.date <= end)
    if branch_id:
        q = q.filter(col == branch_id)
    rows = q.all()

    name_map = {(r.item_type, r.item_id): r.item_name for r in rows}

    med_ids = [iid for (t, iid) in name_map if t == "medicine"]
    if med_ids:
        for m in db.query(DBMedicine.id, DBMedicine.name).filter(
            DBMedicine.id.in_(med_ids)
        ):
            name_map[("medicine", m.id)] = m.name
    dev_ids = [iid for (t, iid) in name_map if t == "medical_device"]
    if dev_ids:
        for d in db.query(DBMedicalDevice.id, DBMedicalDevice.name).filter(
            DBMedicalDevice.id.in_(dev_ids)
        ):
            name_map[("medical_device", d.id)] = d.name

    json_rows: list[dict] = []
    for r in rows:
        dt_iso = r.date.isoformat() if r.date else ""
        name = name_map.get((r.item_type, r.item_id), r.item_name)
        item = {"type": r.item_type, "name": name, "quantity": r.quantity}
        json_rows.append({"id": r.id, "datetime": dt_iso, "items": [item]})

    return {"data": json_rows}


@app.get("/reports/arrivals")
async def get_arrivals_report(
    request: Request,
    date_from: str,
    date_to: str,
    branch_id: str | None = Query(None),
    export: str | None = Query(None),
    format: str | None = Query(None),
    db: Session = Depends(get_db),
):
    result = await build_arrivals_json_payload(
        branch_id=branch_id, date_from=date_from, date_to=date_to, db=db
    )

    if _wants_excel(export, format):
        rows: list[list[str]] = []
        for r in result.get("data", []):
            items_list = r.get("items", []) or []
            items_human = "; ".join(
                f"{i.get('name','')} — {i.get('quantity','')}" for i in items_list
            )
            rows.append([
                _to_almaty_str(r.get("datetime", "")),
                items_human,
            ])

        content = _render_xlsx(
            headers=[
                "Дата и время",
                "Поступило (наименование — кол-во)",
            ],
            rows=rows,
            sheet_name="Поступления",
        )

        safe_to = date_to or datetime.now(ALMATY_TZ).strftime("%Y-%m-%d")
        return Response(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=_ascii_headers(
                f"arrivals_report_{safe_to}.xlsx"
            ),
        )

    return result



def _parse_date(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            d = datetime.strptime(s, fmt)
            return datetime(d.year, d.month, d.day)
        except ValueError:
            continue
    raise HTTPException(status_code=400, detail=f"Bad date format: {s}")


def pick_col(model, *names, fallback=None):
    """Return the first existing column name from names for SQL use."""
    cols = {str(c.name) for c in model.__table__.columns}
    for n in names:
        if n in cols:
            return n
    return fallback


def _to_almaty_str(dt_value) -> str:
    """
    Accepts: datetime or ISO str. Returns 'YYYY-MM-DD HH:MM:SS' in Asia/Almaty.
    """
    if isinstance(dt_value, str):
        s = dt_value.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(s)
        except Exception:
            return dt_value
    else:
        dt = dt_value

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(ALMATY_TZ).strftime("%Y-%m-%d %H:%M:%S")


def _render_xlsx(headers: list[str], rows: list[list[str]], sheet_name: str = "Sheet1") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for r in rows:
        ws.append(r)
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _ascii_headers(filename_ascii: str) -> dict[str, str]:
    return {"Content-Disposition": f"attachment; filename={filename_ascii}"}


def _parse_ymd(s: str | None, end_of_day: bool = False) -> datetime | None:
    if not s:
        return None
    s = s.strip()
    try:
        d = datetime.strptime(s, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(f"Bad date format: {s}") from exc
    if end_of_day:
        return datetime(
            d.year,
            d.month,
            d.day,
            23,
            59,
            59,
            999999,
            tzinfo=timezone.utc,
        )
    return datetime(d.year, d.month, d.day, tzinfo=timezone.utc)


def _wants_excel(export: str | None, format_: str | None) -> bool:
    e, f = (export or "").lower(), (format_ or "").lower()
    return e in {"excel", "xlsx"} or f in {"excel", "xlsx"}


def _get_main_branch_id(db):
    global MAIN_BRANCH_ID
    if MAIN_BRANCH_ID:
        return MAIN_BRANCH_ID
    row = db.execute(
        text(
            "SELECT id FROM branches WHERE is_main=true ORDER BY created_at ASC LIMIT 1"
        )
    ).first()
    if not row:
        row = db.execute(
            text("SELECT id FROM branches ORDER BY created_at ASC LIMIT 1")
        ).first()
    MAIN_BRANCH_ID = row[0]
    return MAIN_BRANCH_ID


# Backward compatibility for older callers
def fmt_almaty(dt: datetime) -> str:
    return _to_almaty_str(dt)


def render_xlsx(headers: list[str], rows: list[list[str]], sheet_name: str = "Sheet1") -> bytes:
    return _render_xlsx(headers, rows, sheet_name)


def _ascii_download_headers(filename_ascii: str) -> dict[str, str]:
    return _ascii_headers(filename_ascii)

# backward compatibility
ascii_download_headers = _ascii_download_headers


def pick_arrival_branch_col(Arrival):
    for name in ("to_branch_id", "branch_id"):
        if hasattr(Arrival, name):
            return getattr(Arrival, name)
    for name in ("to_branch", "branch"):
        if hasattr(Arrival, name):
            rel = getattr(Arrival, name)
            return rel.id
    raise RuntimeError("Arrival: no branch column found")


def build_wh_stock_json(
    db, branch_id: str, date_from: str | None, date_to: str | None
) -> dict:
    sql_current = """
        SELECT 'medicine' AS item_type, m.id AS item_id, m.name AS name,
               COALESCE(c.name, '—') AS category, m.quantity AS quantity
        FROM medicines m
        LEFT JOIN categories c ON c.id = m.category_id
        WHERE m.branch_id = :b AND m.quantity > 0
        UNION ALL
        SELECT 'medical_device' AS item_type, d.id AS item_id, d.name AS name,
               COALESCE(c.name, '—') AS category, d.quantity AS quantity
        FROM medical_devices d
        LEFT JOIN categories c ON c.id = d.category_id
        WHERE d.branch_id = :b AND d.quantity > 0
        ORDER BY item_type, name
    """

    start = None
    end = None
    if date_from and date_to:
        start = datetime.fromisoformat(date_from)
        end = datetime.fromisoformat(date_to)

    if not (start and end):
        rows = db.execute(text(sql_current), {"b": branch_id}).mappings().all()
        return {"data": rows}

    arrival_branch_col = pick_col(DBArrival, "to_branch_id", "branch_id", "branch")
    if not arrival_branch_col:
        rows = db.execute(text(sql_current), {"b": branch_id}).mappings().all()
        return {"data": rows}

    sql_dated = f"""
        WITH start_in AS (
            SELECT item_type, item_id, COALESCE(SUM(quantity),0) AS qty
            FROM arrivals
            WHERE {arrival_branch_col} = :b AND date < :start
            GROUP BY item_type, item_id
        ),
        start_out AS (
            SELECT di.item_type, di.item_id, COALESCE(SUM(di.quantity),0) AS qty
            FROM dispensing_items di
            JOIN dispensing_records dr ON dr.id = di.record_id
            WHERE dr.branch_id = :b AND dr.date < :start
            GROUP BY di.item_type, di.item_id
        ),
        pin AS (
            SELECT item_type, item_id, COALESCE(SUM(quantity),0) AS qty
            FROM arrivals
            WHERE {arrival_branch_col} = :b AND date >= :start AND date <= :end
            GROUP BY item_type, item_id
        ),
        pout AS (
            SELECT di.item_type, di.item_id, COALESCE(SUM(di.quantity),0) AS qty
            FROM dispensing_items di
            JOIN dispensing_records dr ON dr.id = di.record_id
            WHERE dr.branch_id = :b AND dr.date >= :start AND dr.date <= :end
            GROUP BY di.item_type, di.item_id
        ),
        base AS (
            SELECT item_type, item_id, COALESCE(SUM(qty),0) AS quantity
            FROM (
                SELECT item_type, item_id, qty FROM start_in
                UNION ALL
                SELECT item_type, item_id, -qty FROM start_out
                UNION ALL
                SELECT item_type, item_id, qty FROM pin
                UNION ALL
                SELECT item_type, item_id, -qty FROM pout
            ) s
            GROUP BY item_type, item_id
        )
        SELECT b.item_type, b.item_id, b.quantity,
               COALESCE(m.name, d.name) AS name,
               COALESCE(mc.name, dc.name, '—') AS category
        FROM base b
        LEFT JOIN medicines m ON (b.item_type = 'medicine' AND m.id = b.item_id)
        LEFT JOIN categories mc ON mc.id = m.category_id
        LEFT JOIN medical_devices d ON (b.item_type = 'medical_device' AND d.id = b.item_id)
        LEFT JOIN categories dc ON dc.id = d.category_id
        WHERE b.quantity > 0
        ORDER BY b.item_type, name
    """
    rows = db.execute(
        text(sql_dated), {"b": branch_id, "start": start, "end": end}
    ).mappings().all()
    return {"data": rows}


def build_wh_arrivals_json(
    db, start: datetime | None, end: datetime | None
) -> dict:
    q = db.query(DBArrival)

    date_col = getattr(DBArrival, "created_at", None)
    if date_col is None:
        date_col = getattr(DBArrival, "date")
    date_attr = date_col.key

    if start:
        q = q.filter(date_col >= start)
    if end:
        q = q.filter(date_col <= end)

    if hasattr(DBArrival, "items"):
        q = q.options(joinedload(DBArrival.items))

    rows = q.order_by(date_col.asc()).all()

    json_rows: list[dict] = []
    for r in rows:
        dt = getattr(r, date_attr, None)
        dt_iso = dt.isoformat() if dt else ""
        if hasattr(r, "items"):
            items = [
                {
                    "type": it.item_type,
                    "name": it.item_name,
                    "quantity": it.quantity,
                }
                for it in (r.items or [])
            ]
        else:
            items = [
                {
                    "type": getattr(r, "item_type", ""),
                    "name": getattr(r, "item_name", ""),
                    "quantity": getattr(r, "quantity", 0),
                }
            ]
        json_rows.append({"id": r.id, "datetime": dt_iso, "items": items})

    return {"data": json_rows}


def build_wh_dispatches_json(
    db, start: datetime | None, end: datetime | None
) -> dict:
    q = db.query(DBShipment)

    date_col = getattr(DBShipment, "created_at")
    date_attr = date_col.key

    if start:
        q = q.filter(date_col >= start)
    if end:
        q = q.filter(date_col <= end)

    rows = q.order_by(date_col.asc()).all()

    json_rows: list[dict] = []
    for r in rows:
        items = (
            db.query(DBShipmentItem)
            .filter(DBShipmentItem.shipment_id == r.id)
            .all()
        )
        items_data = [
            {"type": it.item_type, "name": it.item_name, "quantity": it.quantity}
            for it in items
        ]
        dt = getattr(r, date_attr, None)
        dt_iso = dt.isoformat() if dt else ""
        json_rows.append({"id": r.id, "datetime": dt_iso, "items": items_data})

    return {"data": json_rows}


@app.get("/admin/warehouse/reports/stock")
def admin_warehouse_stock(
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    export: str | None = Query(None),
    format: str | None = Query(None),
    db: Session = Depends(get_db),
):
    """
    Warehouse stock report (no branches involved):
    Sum arrivals by (item_type, item_id) within the selected period.
    JSON by default; if export=excel|xlsx — return an XLSX download.
    """
    try:
        start = _parse_ymd(date_from)
        end = _parse_ymd(date_to, end_of_day=True)

        date_col = getattr(DBArrival, "created_at", None)
        if date_col is None:
            date_col = getattr(DBArrival, "date")

        q = (
            select(
                getattr(DBArrival, "item_type").label("item_type"),
                getattr(DBArrival, "item_id").label("item_id"),
                func.coalesce(func.sum(getattr(DBArrival, "quantity")), 0).label("qty"),
            ).group_by(getattr(DBArrival, "item_type"), getattr(DBArrival, "item_id"))
        )

        if start:
            q = q.where(date_col >= start)
        if end:
            q = q.where(date_col <= end)

        rows = db.execute(q).all()

        result = []
        for item_type, item_id, qty in rows:
            if not qty or qty <= 0:
                continue

            name = None
            category = None

            if item_type == "medicine":
                m = db.get(DBMedicine, item_id)
                if m:
                    name = getattr(m, "name", None)
                    cat_id = getattr(m, "category_id", None)
                    if cat_id:
                        cat = db.get(DBCategory, cat_id)
                        category = getattr(cat, "name", None)
            elif item_type == "medical_device":
                md = db.get(DBMedicalDevice, item_id)
                if md:
                    name = getattr(md, "name", None)
                    cat_id = getattr(md, "category_id", None)
                    if cat_id:
                        cat = db.get(DBCategory, cat_id)
                        category = getattr(cat, "name", None)

            result.append(
                {
                    "name": name or "-",
                    "category": category or "—",
                    "quantity": int(qty),
                    "item_type": item_type,
                    "item_id": str(item_id),
                }
            )

        if ((export or "").lower() in {"excel", "xlsx"}) or (
            (format or "").lower() in {"excel", "xlsx"}
        ):
            rows_x = [[r["name"], r["category"], r["quantity"]] for r in result]
            content = _render_xlsx(
                ["Наименование", "Категория", "Количество"],
                rows_x,
                "Остатки",
            )
            safe_to = date_to or "today"
            return Response(
                content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=_ascii_download_headers(f"warehouse_stock_{safe_to}.xlsx"),
            )

        return {"data": result}

    except Exception as e:
        log.exception("Warehouse stock report failed")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/admin/warehouse/reports/arrivals")
def admin_wh_arrivals(
    request: Request,
    branch_id: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    export: Optional[str] = Query(None),
    format: Optional[str] = Query(None),
):
    try:
        start = _parse_ymd(date_from)
        end = _parse_ymd(date_to, end_of_day=True)
        with SessionLocal() as db:
            payload = build_wh_arrivals_json(db, start, end)
            if _wants_excel(export, format):
                rows: list[list[str]] = []
                for r in payload.get("data", []):
                    items = "; ".join(
                        f"{i.get('name','')} — {i.get('quantity','')}" for i in r.get("items", [])
                    )
                    rows.append([_to_almaty_str(r.get("datetime", "")), items])
                content = _render_xlsx(
                    ["Дата и время", "Поступило (наименование — кол-во)"],
                    rows,
                    "Поступления",
                )
                safe_to = date_to or datetime.now(ALMATY_TZ).strftime("%Y-%m-%d")
                return Response(
                    content,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=_ascii_download_headers(f"warehouse_arrivals_{safe_to}.xlsx"),
                )
            return payload
    except Exception as e:
        logger.exception("admin_wh_arrivals error")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/admin/warehouse/reports/dispatches")
def admin_wh_dispatches(
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    export: Optional[str] = Query(None),
    format: Optional[str] = Query(None),
):
    try:
        start = _parse_ymd(date_from)
        end = _parse_ymd(date_to, end_of_day=True)
        with SessionLocal() as db:
            payload = build_wh_dispatches_json(db, start, end)
            if _wants_excel(export, format):
                rows: list[list[str]] = []
                for r in payload.get("data", []):
                    items = "; ".join(
                        f"{i.get('name','')} — {i.get('quantity','')}" for i in r.get("items", [])
                    )
                    rows.append([_to_almaty_str(r.get("datetime", "")), items])
                content = _render_xlsx(
                    ["Дата и время", "Отправлено (наименование — кол-во)"],
                    rows,
                    "Отправки",
                )
                safe_to = date_to or datetime.now(ALMATY_TZ).strftime("%Y-%m-%d")
                return Response(
                    content,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers=_ascii_download_headers(
                        f"warehouse_dispatches_{safe_to}.xlsx"
                    ),
                )
            return payload
    except Exception as e:
        logger.exception("admin_wh_dispatches error")
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/reports/stock")
def get_stock_report(
    branch_id: str,
    date_from: str | None = None,
    date_to: str | None = None,
):
    # parse dates if provided
    start = None
    end = None
    if date_from and date_to:
        from datetime import datetime

        start = datetime.fromisoformat(date_from)
        end = datetime.fromisoformat(date_to)

    with SessionLocal() as db:
        # === PATH A: no date range => current on-hand ===
        if not (start and end):
            sql = """
                SELECT 'medicine' AS item_type, m.id AS item_id, m.name AS name,
                       COALESCE(c.name, '—') AS category, m.quantity AS quantity
                FROM medicines m
                LEFT JOIN categories c ON c.id = m.category_id
                WHERE m.branch_id = :b AND m.quantity > 0
                UNION ALL
                SELECT 'medical_device' AS item_type, d.id AS item_id, d.name AS name,
                       COALESCE(c.name, '—') AS category, d.quantity AS quantity
                FROM medical_devices d
                LEFT JOIN categories c ON c.id = d.category_id
                WHERE d.branch_id = :b AND d.quantity > 0
                ORDER BY item_type, name
            """
            rows = db.execute(text(sql), {"b": branch_id}).mappings().all()
            return {"data": rows}

        # === PATH B: date range present => closing balance as of `end` ===
        arrival_branch_col = pick_col(DBArrival, "to_branch_id", "branch_id", "branch")
        if not arrival_branch_col:
            sql_fallback = """
                SELECT 'medicine' AS item_type, m.id AS item_id, m.name AS name,
                       COALESCE(c.name, '—') AS category, m.quantity AS quantity
                FROM medicines m
                LEFT JOIN categories c ON c.id = m.category_id
                WHERE m.branch_id = :b AND m.quantity > 0
                UNION ALL
                SELECT 'medical_device' AS item_type, d.id AS item_id, d.name AS name,
                       COALESCE(c.name, '—') AS category, d.quantity AS quantity
                FROM medical_devices d
                LEFT JOIN categories c ON c.id = d.category_id
                WHERE d.branch_id = :b AND d.quantity > 0
                ORDER BY item_type, name
            """
            rows = db.execute(text(sql_fallback), {"b": branch_id}).mappings().all()
            return {"data": rows}

        sql_dated = f"""
            WITH start_in AS (
                SELECT item_type, item_id, COALESCE(SUM(quantity),0) AS qty
                FROM arrivals
                WHERE {arrival_branch_col} = :b AND date < :start
                GROUP BY item_type, item_id
            ),
            start_out AS (
                SELECT di.item_type, di.item_id, COALESCE(SUM(di.quantity),0) AS qty
                FROM dispensing_items di
                JOIN dispensing_records dr ON dr.id = di.record_id
                WHERE dr.branch_id = :b AND dr.date < :start
                GROUP BY di.item_type, di.item_id
            ),
            pin AS (
                SELECT item_type, item_id, COALESCE(SUM(quantity),0) AS qty
                FROM arrivals
                WHERE {arrival_branch_col} = :b AND date >= :start AND date <= :end
                GROUP BY item_type, item_id
            ),
            pout AS (
                SELECT di.item_type, di.item_id, COALESCE(SUM(di.quantity),0) AS qty
                FROM dispensing_items di
                JOIN dispensing_records dr ON dr.id = di.record_id
                WHERE dr.branch_id = :b AND dr.date >= :start AND dr.date <= :end
                GROUP BY di.item_type, di.item_id
            ),
            base AS (
                SELECT item_type, item_id, COALESCE(SUM(qty),0) AS quantity
                FROM (
                    SELECT item_type, item_id, qty FROM start_in
                    UNION ALL
                    SELECT item_type, item_id, -qty FROM start_out
                    UNION ALL
                    SELECT item_type, item_id, qty FROM pin
                    UNION ALL
                    SELECT item_type, item_id, -qty FROM pout
                ) s
                GROUP BY item_type, item_id
            )
            SELECT b.item_type, b.item_id, b.quantity,
                   COALESCE(m.name, d.name) AS name,
                   COALESCE(mc.name, dc.name, '—') AS category
            FROM base b
            LEFT JOIN medicines m ON (b.item_type = 'medicine' AND m.id = b.item_id)
            LEFT JOIN categories mc ON mc.id = m.category_id
            LEFT JOIN medical_devices d ON (b.item_type = 'medical_device' AND d.id = b.item_id)
            LEFT JOIN categories dc ON dc.id = d.category_id
            WHERE b.quantity > 0
            ORDER BY b.item_type, name
        """
        rows = db.execute(
            text(sql_dated), {"b": branch_id, "start": start, "end": end}
        ).mappings().all()
        return {"data": rows}


@app.get("/reports/stock/export")
def export_stock_xlsx(
    branch_id: str = Query(...),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
):
    if Workbook is None:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    payload = get_stock_report(
        branch_id=branch_id, date_from=date_from, date_to=date_to
    )
    records = payload.get("data", [])
    rows = [
        {
            "Название": r["name"],
            "Категория": r.get("category", ""),
            "Количество": r["quantity"],
        }
        for r in records
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"
    headers = ["Название", "Категория", "Количество"]
    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "Остатки.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/reports/stock/item_details")
async def get_stock_item_details(
    branch_id: str,
    type: str,
    item_id: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
):
    try:
        tz = ZoneInfo("Asia/Almaty")
        if date_to:
            dt_to = datetime.fromisoformat(date_to)
            if dt_to.tzinfo is None:
                dt_to = dt_to.replace(tzinfo=tz)
            else:
                dt_to = dt_to.astimezone(tz)
            dt_to = dt_to.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:
            now = datetime.now(tz)
            dt_to = now.replace(hour=23, minute=59, second=59, microsecond=999999)

        if date_from:
            dt_from = datetime.fromisoformat(date_from)
            if dt_from.tzinfo is None:
                dt_from = dt_from.replace(tzinfo=tz)
            else:
                dt_from = dt_from.astimezone(tz)
            dt_from = dt_from.replace(hour=0, minute=0, second=0, microsecond=0)
            start_utc = dt_from.astimezone(timezone.utc).replace(tzinfo=None)
        else:
            start_utc = None
        end_utc = dt_to.astimezone(timezone.utc).replace(tzinfo=None)

        params = {"b": branch_id, "t": type, "i": item_id, "end": end_utc}
        if start_utc:
            params["start"] = start_utc

        sql_in = (
            "SELECT date, quantity FROM arrivals "
            "WHERE branch_id = :b AND item_type = :t AND item_id = :i AND date <= :end"
        )
        if start_utc:
            sql_in += " AND date >= :start"
        sql_in += " ORDER BY date"
        incoming_rows = db.execute(text(sql_in), params).fetchall()
        incoming = [
            {"date": r[0].isoformat(), "qty": int(r[1])} for r in incoming_rows
        ]

        sql_out = (
            "SELECT dr.date, di.quantity FROM dispensing_records dr "
            "JOIN dispensing_items di ON dr.id = di.record_id "
            "WHERE dr.branch_id = :b AND di.item_type = :t AND di.item_id = :i AND dr.date <= :end"
        )
        if start_utc:
            sql_out += " AND dr.date >= :start"
        sql_out += " ORDER BY dr.date"
        outgoing_rows = db.execute(text(sql_out), params).fetchall()
        outgoing = [
            {"date": r[0].isoformat(), "qty": int(r[1])} for r in outgoing_rows
        ]

        total_in = sum(r["qty"] for r in incoming)
        total_out = sum(r["qty"] for r in outgoing)

        return {
            "incoming": incoming,
            "outgoing": outgoing,
            "total_in": total_in,
            "total_out": total_out,
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Calendar endpoints
@app.get("/calendar/dispensing")
async def get_calendar_dispensing(
    start: Optional[str] = None,
    end: Optional[str] = None,
    date: Optional[str] = None,
    branch_id: Optional[str] = None,
    aggregate: Optional[int] = None,
    # legacy params
    month: Optional[str] = None,
    patient_id: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Calendar dispensing endpoint supporting summary and day listing."""
    try:
        tz = ZoneInfo("Asia/Almaty")

        # Aggregate mode: monthly summary
        if aggregate == 1 and start and end:
            start_local = datetime.strptime(start, "%Y-%m-%d").replace(
                tzinfo=tz, hour=0, minute=0, second=0, microsecond=0
            )
            end_local = datetime.strptime(end, "%Y-%m-%d").replace(
                tzinfo=tz, hour=0, minute=0, second=0, microsecond=0
            ) + timedelta(days=1)
            start_utc = start_local.astimezone(timezone.utc).replace(tzinfo=None)
            end_utc = end_local.astimezone(timezone.utc).replace(tzinfo=None)
            query = db.query(DBDispensingRecord).filter(
                DBDispensingRecord.date >= start_utc,
                DBDispensingRecord.date < end_utc,
            )
            if branch_id:
                query = query.filter(DBDispensingRecord.branch_id == branch_id)
            records = query.all()
            counts: dict[str, int] = {}
            for rec in records:
                dt = rec.date
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                local_dt = dt.astimezone(tz)
                key = local_dt.strftime("%Y-%m-%d")
                counts[key] = counts.get(key, 0) + 1
            data = [
                {"date": d, "count": counts[d]} for d in sorted(counts.keys())
            ]
            return {"data": data}

        # Day list mode
        if date:
            day_local = datetime.strptime(date, "%Y-%m-%d").replace(
                tzinfo=tz, hour=0, minute=0, second=0, microsecond=0
            )
            next_day_local = day_local + timedelta(days=1)
            start_utc = day_local.astimezone(timezone.utc).replace(tzinfo=None)
            end_utc = next_day_local.astimezone(timezone.utc).replace(tzinfo=None)
            query = (
                db.query(DBDispensingRecord, DBBranch)
                .join(DBBranch, DBDispensingRecord.branch_id == DBBranch.id)
                .filter(
                    DBDispensingRecord.date >= start_utc,
                    DBDispensingRecord.date < end_utc,
                )
            )
            if branch_id:
                query = query.filter(DBDispensingRecord.branch_id == branch_id)
            records = query.order_by(DBDispensingRecord.date.asc()).all()
            result = []
            for rec, branch in records:
                dt = rec.date
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                local_dt = dt.astimezone(tz)
                result.append(
                    {
                        "id": rec.id,
                        "time": local_dt.strftime("%H:%M:%S"),
                        "patient_name": rec.patient_name,
                        "employee_name": rec.employee_name,
                        "branch_name": branch.name if branch else "",
                    }
                )
            return {"data": result}

        # Legacy behaviour for branch calendar
        if not (branch_id and month):
            raise HTTPException(status_code=400, detail="Missing parameters")
        start_local = datetime.strptime(month, "%Y-%m").replace(
            tzinfo=tz, day=1, hour=0, minute=0, second=0, microsecond=0
        )
        if start_local.month == 12:
            end_local = start_local.replace(year=start_local.year + 1, month=1)
        else:
            end_local = start_local.replace(month=start_local.month + 1)
        start_utc = start_local.astimezone(timezone.utc).replace(tzinfo=None)
        end_utc = end_local.astimezone(timezone.utc).replace(tzinfo=None)

        query = db.query(DBDispensingRecord).filter(
            DBDispensingRecord.branch_id == branch_id,
            DBDispensingRecord.date >= start_utc,
            DBDispensingRecord.date < end_utc,
        )
        if patient_id:
            query = query.filter(DBDispensingRecord.patient_id == patient_id)

        records = query.all()
        calendar_data = {}
        for record in records:
            record_dt = record.date
            if record_dt.tzinfo is None:
                record_dt = record_dt.replace(tzinfo=timezone.utc)
            local_dt = record_dt.astimezone(tz)
            day_key = str(local_dt.day)
            if day_key not in calendar_data:
                calendar_data[day_key] = []
            calendar_data[day_key].append(
                {
                    "id": record.id,
                    "created_at": record.date.isoformat(),
                    "patient_id": record.patient_id,
                    "employee_name": record.employee_name,
                }
            )

        return {"data": calendar_data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/calendar/dispensing/day")
async def get_calendar_dispensing_day(
    branch_id: str,
    patient_id: str,
    date: str,
    db: Session = Depends(get_db),
):
    """Return dispensing details for a specific patient on a given day."""
    try:
        tz = ZoneInfo("Asia/Almaty")
        day_local = datetime.strptime(date, "%Y-%m-%d").replace(
            tzinfo=tz, hour=0, minute=0, second=0, microsecond=0
        )
        next_day_local = day_local + timedelta(days=1)
        start_utc = day_local.astimezone(timezone.utc).replace(tzinfo=None)
        end_utc = next_day_local.astimezone(timezone.utc).replace(tzinfo=None)

        records = (
            db.query(DBDispensingRecord)
            .options(joinedload(DBDispensingRecord.items))
            .filter(
                DBDispensingRecord.branch_id == branch_id,
                DBDispensingRecord.patient_id == patient_id,
                DBDispensingRecord.date >= start_utc,
                DBDispensingRecord.date < end_utc,
            )
            .order_by(DBDispensingRecord.date.asc())
            .all()
        )

        result = []
        for record in records:
            record_dt = record.date
            if record_dt.tzinfo is None:
                record_dt = record_dt.replace(tzinfo=timezone.utc)
            local_dt = record_dt.astimezone(tz)
            result.append(
                {
                    "time": local_dt.strftime("%H:%M"),
                    "employee_name": record.employee_name,
                    "items": [
                        {
                            "type": item.item_type,
                            "name": item.item_name,
                            "quantity": item.quantity,
                        }
                        for item in record.items
                    ],
                }
            )

        return {"data": result}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
